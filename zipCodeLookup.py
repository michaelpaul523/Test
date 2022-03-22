import re, openpyxl, os, datetime
from openpyxl.styles import Alignment, Color, PatternFill, Font, colors
from openpyxl.styles.borders import Border, Side
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException, NoSuchElementException

#Font, Border, and Fill variables
yellowFill = PatternFill(start_color = '00FFFF00', end_color = '00FFFF00', fill_type = 'solid')
greyFill = PatternFill(start_color = '00D4D4D4', end_color = '00D4D4D4', fill_type = 'solid')
thinBorder = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
boldFont = Font(name='Calibri', size=11, bold=True, color=colors.BLACK)
normalFont = Font(name='Calibri', size=11, bold=False, color=colors.BLACK)
centerAlign = Alignment(horizontal='center')

workbook = openpyxl.load_workbook(r'C:\Users\michael.dickinson\Desktop\ZipCodeLookup\zipCodeTemplate.xlsx')
sheet = workbook['Sheet1']
regex=re.compile(r'-\d{4}')

while True:
	while True:
		low=input('Enter low range: ')
		if low.isdigit() == False: print('Please enter a number.')
		else: break
	while True:
		high=input('Enter high range: ')
		if high.isdigit() == False: print('Please enter a number.')
		else: break
	if int(low) >= int(high): print('Low range bust be greater than high range.')
	else:break

while True:
	oddEven=input('Odd, even or both? (Enter an o, e or b.) ').lower()
	if oddEven not in ['o','e','b']: print('Please enter an o, e or b.')
	else: break

low,high=int(low),int(high)
if oddEven == 'o' and low%2 != 1: low-=1
if oddEven == 'e' and low%2 != 0: low-=1
if oddEven == 'o' and high%2 != 1: high+=1
if oddEven == 'e' and high%2 != 0: high+=1
if oddEven == 'o':addressNums = range(low,high+1,2)
elif oddEven == 'e':addressNums = range(low,high+1,2)
elif oddEven == 'b': addressNums= range(low,high+1)

roadName=input('Enter road name: ').upper()
postalCommunity=input('Enter Postal Community: ').upper()
while True:
	zipCode=input('Enter ZIP Code: ' )
	if zipCode.isdigit() == False or len(zipCode) != 5: print('Please enter a 5 digit number.')
	else: break

browser = webdriver.Firefox()
browser.get(r'https://tools.usps.com/zip-code-lookup.htm?byaddress')

addressDict={}
for i in addressNums:
	try: myElem = WebDriverWait(browser,3).until(EC.presence_of_element_located((By.XPATH, "//*[@id='tAddress']")))
	except TimeoutException: pass
	addressField=browser.find_element_by_xpath("//*[@id='tAddress']")
	addressField.send_keys(str(i) + ' ' + roadName)
	cityField=browser.find_element_by_xpath("//*[@id='tCity']")
	cityField.send_keys(postalCommunity)
	stateField=Select(browser.find_element_by_xpath("//*[@id='tState']"))
	stateField.select_by_value('AL')
	zipField=browser.find_element_by_xpath("//*[@id='tZip-byaddress']")
	zipField.send_keys(zipCode)
	findButton=browser.find_element_by_xpath("//*[@id='zip-by-address']")
	findButton.click()
	try:
		try: myElem = WebDriverWait(browser,3).until(EC.presence_of_element_located((By.XPATH, "//*[@class='zipcode-result-address']")))
		except TimeoutException: pass
		resultAddress=browser.find_element_by_xpath("//*[@class='zipcode-result-address']")
		resultMO=regex.search(resultAddress.text)
		if resultMO: 
			print('\nMatch Found:')
			print(resultAddress.text)
			addressDict[str(i) + ' ' + roadName]=resultAddress.text
		searchAgain=browser.find_element_by_xpath("//*[@id='search-address-again']")
		searchAgain.click()	
	except NoSuchElementException: browser.refresh()

excelRow=2		
for k,v in addressDict.items():
	sheet.cell(row=excelRow,column=1).value=k
	sheet.cell(row=excelRow,column=2).value=v
	for cell in sheet[str(excelRow):str(excelRow)]:
		cell.alignment=centerAlign
		cell.font=normalFont
		cell.border=thinBorder
	excelRow+=1

#Adjust Column width and re-apply border to columns incase previous border was overwritten
excelColumn = 1
for columnCells in sheet.columns:
	length = 0
	for cell in columnCells:
		if cell.value == None:
			pass
		elif isinstance(cell.value, int):
			pass
		elif len(cell.value) > length:
			length = len(cell.value)
	columnLetter = openpyxl.utils.get_column_letter(excelColumn)
	sheet.column_dimensions[columnLetter].width = length + 3
	sheet.column_dimensions[columnLetter].border = thinBorder
	excelColumn +=1

timestamp=datetime.datetime.now().strftime('%m-%d-%y')
outName=str(zipCode) + ' ' + roadName + ' ' + str(low) + '-' + str(high) + ' ZIP LOOKUP ' + timestamp + '.xlsx'
outFolder=r'C:\Users\michael.dickinson\Desktop\ZipCodeLookup\Completed'
workbook.save(os.path.join(outFolder,outName))
